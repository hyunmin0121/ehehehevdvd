#공지 시스템을 사용하기위해선 nchannel.xlsx 파일이 필요합니다.

import discord
import os
import requests
from bs4 import BeautifulSoup
import datetime
import openpyxl
import asyncio
import os
import sys
import urllib.request
import json
import time
from matplotlib import pyplot as plt

client = discord.Client()
async  def status_task():
    while True:
        game = discord.Game("!코로나 도움말")
        await client.change_presence(status=discord.Status.online, activity=game)
        await asyncio.sleep(10)
        game = discord.Game(str(len(set(client.guilds))) + "개의 서버에서 " + str(len(set(client.get_all_members())))+ "명이 이용중")
        await client.change_presence(status=discord.Status.online, activity=game)
        await asyncio.sleep(10)
        game = discord.Game("코로나 증상 의심시 국번없이 1339로 전화하세요")
        await client.change_presence(status=discord.Status.online, activity=game)
        await asyncio.sleep(10)

async  def news():
    while True:
        request = requests.get('https://www.yna.co.kr/safe/index')
        request.encoding = 'utf-8'
        html = request.text
        soup = BeautifulSoup(html, 'html.parser')
        links = soup.select('#container > div > div.inner > section > div > div.list-wrap > div > ul.list01 > li > article > div.news-con > a > h3')
        link = soup.select('#container > div > div.inner > section > div > div.list-wrap > div > ul.list01 > li > article > div.news-con > a')
        f = open('news.txt', 'r', encoding='UTF8')
        if str(links[0].text).replace('&apos;','')!= f.read():
            print("뉴스 기사 업데이트")
            f.close()
            f = open('news.txt', 'w', encoding='UTF8')
            f.write(str(links[0].text).replace('&apos;',''))
            f.close()
            i=1
            file = openpyxl.load_workbook("nchannel.xlsx")
            sheet = file.active
            while True:
                if sheet["A" + str(i)].value == None:
                    file.save("nchannel.xlsx")
                    break;
                if sheet["B" + str(i)].value == 1:
                    embed = discord.Embed(color=0x50BCDF, title="__실시간 뉴스__")
                    embed.add_field(name="코로나 알림 포럼", value=str(links[0].text).replace('&apos;','')+ "\n" + "https:" + str(link[0]).split('<a href="')[1].split('">')[0])
                    embed.set_footer(text="해당메세지는 코로나 알림 포럼 https://discord.gg/nBUGWPf에서 발송되었습니다.")
                    try:
                        await client.get_channel(int(sheet["A" + str(i)].value)).send(embed=embed)
                    except AttributeError:
                        pass;


                i += 1

        await asyncio.sleep(10)



@client.event
async def on_ready():
    bid = (str(client.user.id))
    print("봇 이름 : " + bid)
    bname = client.user.name
    print("봇이름 : " + bname)
    print("STAR")
    client.loop.create_task(status_task())
    client.loop.create_task(news())

@client.event
async def on_message(message):
    if message.content.startswith("!코로나"):
        #await message.channel.send("질병관리본부 코로나19 사이트 접속 마비로 인해 코로나관련 명령을 제외한 나머지 기능들은 정상적으로 이용하실 수 있으니 이용에 참고하시기 바랍니다. 불편을 드려 죄송합니다.\n자세한정보는 https://discord.gg/nBUGWPf에서 확인해주세요")
        request = requests.get('http://ncov.mohw.go.kr/bdBoardList_Real.do')
        html = request.text
        soup = BeautifulSoup(html, 'html.parser')
        links = soup.select('#content > div > div.bv_content > div > ul > li')
        try:
            gmember = client.guilds
            for s in gmember:
                if s == message.guild:
                    gamout=(len(s.members))
            print("\n호출 서버 : " + str(message.guild) + " 서버규모 : " + str(gamout) + "  사용서버 : " + str(len(set(client.guilds))) + " 사용유저 : " + str(len(set(client.get_all_members()))) + "\n" +
                                                 str(links[0].text).replace('(','').replace(')',' :')+ " " +
                                                 str(links[1].text).replace('(','').replace(')',' :')+ " " +
                                                 str(links[2].text).replace('(','').replace(')',' :')+ " " +
                                                 str(links[3].text).replace('(','').replace(')',' :'))
        except:
            pass;
        try:
            if str(message.content).split(' ')[1]==("전체"):
                embed = discord.Embed(color=0xff4040, title="코로나 전체 현황")
                for i in links:
                    embed.add_field(name=str(i.text).split(')')[0].replace('(', ''),value=str(i.text).replace(str(i.text).split(')')[0],''),inline=False)
                sent = await message.channel.send(embed=embed)
                await asyncio.sleep(30)
                await sent.delete()
                await message.delete()

            if str(message.content).split(' ')[1] == ("확진"):
                embed = discord.Embed(color=0xff4040, title="코로나 확진자 현황")
                embed.add_field(name=str(links[0].text).split(')')[0].replace('(', ''),value=str(links[0].text).replace(str(links[0].text).split(')')[0], '').replace(')',''), inline=False)
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("완치"):
                embed = discord.Embed(color=0xff4040, title="코로나 완치자 현황")
                embed.add_field(name=str(links[1].text).split(')')[0].replace('(', ''),value=str(links[1].text).replace(str(links[1].text).split(')')[0], '').replace(')',''), inline=False)
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("사망"):
                embed = discord.Embed(color=0xff4040, title="코로나 사망자 현황")
                embed.add_field(name=str(links[2].text).split(')')[0].replace('(', ''),value=str(links[2].text).replace(str(links[2].text).split(')')[0], '').replace(')',''), inline=False)
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("검사"):
                embed = discord.Embed(color=0xff4040, title="코로나 검사 진행자 현황")
                embed.add_field(name=str(links[3].text).split(')')[0].replace('(', ''),value=str(links[3].text).replace(str(links[3].text).split(')')[0], '').replace(')',''), inline=False)
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("글로벌"):
                embed = discord.Embed(color=0xff4040, title="코로나 국외 확진자 현황")
                embed.add_field(name=str(links[4].text).split(')')[0].replace('(', ''),value=str(links[4].text).replace(str(links[4].text).split(')')[0], '').replace(')','').replace('(',' '), inline=False)
                embed.add_field(name=str(links[5].text).split(')')[0].replace('(', ''),
                                value=str(links[5].text).replace(str(links[5].text).split(')')[0], '').replace(')','').replace('(',' '), inline=False)
                embed.add_field(name=str(links[6].text).split(')')[0].replace('(', ''),
                                value=str(links[6].text).replace(str(links[6].text).split(')')[0], '').replace(')','').replace('(',' '), inline=False)
                embed.add_field(name=str(links[7].text).split(')')[0].replace('(', ''),
                                value=str(links[7].text).replace(str(links[7].text).split(')')[0], '').replace(')','').replace('(',' '), inline=False)
                embed.add_field(name=str(links[8].text).split(')')[0].replace('(', ''),
                                value=str(links[8].text).replace(str(links[8].text).split(')')[0], '').replace(')','').replace('(',' '), inline=False)
                embed.add_field(name=str(links[9].text).split(')')[0].replace('(', ''),
                                value=str(links[9].text).replace(str(links[9].text).split(')')[0], '').replace(')','').replace('(',' '), inline=False)
                embed.add_field(name=str(links[10].text).split(')')[0].replace('(', ''),
                                value=str(links[10].text).replace(str(links[10].text).split(')')[0], '').replace(')','').replace('(',' '), inline=False)
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("예방수칙"):
                embed = discord.Embed(color=0xff4040, title="코로나 예방수칙")
                embed.set_image(url='http://akfh278.dothome.co.kr/image/we.png')
                embed.add_field(name="(자막) 코로나바이러스감염증-19 국민 예방수칙", value="https://www.youtube.com/watch?v=St84msZA5CQ")
                #https://www.youtube.com/watch?v=St84msZA5CQ
                sent = await message.channel.send(embed=embed)
                await asyncio.sleep(15)
                await sent.delete()
                await message.delete()

            if str(message.content).split(' ')[1] == ("행동수칙"):
                embed = discord.Embed(color=0xff4040, title="코로나 행동수칙")
                embed.set_image(url='http://akfh278.dothome.co.kr/image/ha.png')
                sent = await message.channel.send(embed=embed)
                await asyncio.sleep(15)
                await sent.delete()
                await message.delete()

            if str(message.content).split(' ')[1] == ("진료소"):
                jrequest = requests.get('http://www.mohw.go.kr/react/popup_200128.html')
                jrequest.encoding = 'utf-8'
                jhtml = jrequest.text
                jsoup = BeautifulSoup(jhtml, 'html.parser')
                jlinks = jsoup.select('#user-table > tbody > tr')
                resultlist = []
                for s in jlinks:
                    if str(message.content).split('진료소 ')[1] in str(s):
                        resultlist.append(s)
                embed = discord.Embed(color=0xff4040, title="코로나 진료소" + " " + str(message.content).split('진료소 ')[1] + " 검색결과")
                rcount=0
                for i in resultlist:
                    rcount=rcount+1
                    embed.add_field(name=str(message.content).split('진료소 ')[1] + " " +str(rcount) + "번 진료소", value=str(i).split('</th>')[1].replace('<td>','').replace('</td>',' ').replace('<i class="ico_ntc">','').replace('*<span class="hdn">','').replace('</tr>','').replace('</span></i>',''), inline=False)
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("증상"):
                embed = discord.Embed(color=0xff4040, title="코로나바이러스-19 증상 및 행동안내")
                embed.add_field(name="의심 증상",value="고온의 발열, 폐렴 증상,\n기침, 가래, 인후통 등의 호흡기 증상 등\n",inline=False)
                embed.add_field(name="행동 방법", value="해당 증상이 나타나는 경우 병원이나 응급실로 __지양__하시고 선별 진료소를 통해 진료를 받으시길 바랍니다. \n\n코로나감염이 의심된다면 지역보건소 또는 1339를 통해 상담을 받으시기 바랍니다.",inline=False)
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("타임라인"):
                f = open('tlurl.txt', 'r')
                embed = discord.Embed(color=0xff4040, title="실시간 코로나 타임라인")
                embed.set_image(url=f.read())
                f.close()
                await message.channel.send(embed=embed)

            if str(message.content).split(' ')[1] == ("도움말"):
                embed = discord.Embed(color=0xff4040, title="코로나 알림 명령어")
                embed.add_field(name="명령어", value="!코로나 - 국내 코로나 현황을 출력합니다.\n!코로나 글로벌 - 국외 코로나 현황을 출력합니다."
                                                  "\n!코로나 전체 - 국내외 코로나 현황을 출력합니다.\n!코로나 지역 지역명 - 지역명 지역의 코로나 현황을 출력합니다.\n!코로나 확진 - 국내 코로나 확진자 현황을 출력합니다.\n!코로나 완치 - 국내 코로나 완치자 현황을 출력합니다."
                                                  "\n!코로나 사망 - 국내 코로나 사망자 현황을 출력합니다.\n!코로나 검사 - 국내 코로나 검사 진행자 현황을 출력합니다.\n!코로나 동선 - 코로나 동선 목록을 출력합니다.\n!코로나 동선 목록번호 - 목록변수 확진자 동선을 출력 예) !코로나 동선 1",inline=False)
                embed.add_field(name="ETC",value="!코로나 채널설정 - 코로나 알림봇의 공지 혹은 실시간뉴스를 받을 채널로 설정 혹은 해제 합니다.\n!코로나 예방수칙 - 코로나 개인 예방수칙을 전송합니다.\n!코로나 행동수칙 - 코로나 개인 행동수칙을 전송합니다.\n"
                                                 "!코로나 진료소 지역 - 지역의 코로나 진료소를 검색합니다. 예) !코로나 진료소 강남구\n!코로나 증상 - 코로나 증상 및 행동수칙을 안내합니다.\n"
                                                 "!코로나 타임라인 - 코로나 타임라인을 확인합니다.\n\n봇이 관리자권한을 소유하고 있지 않으면 일부기능이 사용이 제한 될수있습니다.", inline=False)
                #http://ncov.mohw.go.kr/front_new/modules/img_view.jsp?img_loc=/upload/mwEditor/202002/1582512305707_20200224114505.jpg
                #http://ncov.mohw.go.kr/front_new/modules/img_view.jsp?img_loc=/upload/mwEditor/202002/1582512317532_20200224114517.jpg
                embed.add_field(name="코로나 알림 포럼", value="https://discord.gg/nBUGWPf")
                embed.add_field(name="코로나알림 봇초대", value=" http://bitly.kr/xcdqKI1P",inline=True)
                embed.set_footer(text="")
                sent = await message.channel.send(embed=embed)
                await asyncio.sleep(30)
                await sent.delete()
                await message.delete()

            if str(message.content).split(' ')[1] == ("채널설정"):
                if message.author.guild_permissions.administrator:
                    file = openpyxl.load_workbook("nchannel.xlsx")
                    sheet = file.active
                    i=1
                    while True:
                        if sheet["A" + str(i)].value == str(message.channel.id):
                            if sheet["B" + str(i)].value == 1:
                                sheet["B" + str(i)].value = 0
                                await message.channel.send(str(message.channel) + "채널이 코로나 알림봇 채널에서 제외되었습니다")
                            else:
                                sheet["B" + str(i)].value = 1
                                await message.channel.send(str(message.channel) + "채널이 코로나 알림봇 채널로 설정되었습니다")
                            file.save("nchannel.xlsx")
                            break

                        if sheet["A" + str(i)].value == None:
                            sheet["A" + str(i)].value = str(message.channel.id)
                            sheet["B" + str(i)].value = 1
                            await message.channel.send(str(message.channel) + "채널이 코로나 알림봇 채널로 설정되었습니다")
                            file.save("nchannel.xlsx")
                            break
                        i += 1
                else:
                    await message.channel.send("관리자 권한을 소유하고있지 않습니다.")

            if str(message.content).split(' ')[1] == ("동선"):
                request = requests.get('https://www.yna.co.kr/view/AKR20200221051400505?section=safe/news')
                request.encoding = 'utf-8'
                html = request.text
                soup = BeautifulSoup(html, 'html.parser')
                links = soup.select('#articleWrap > div > p')
                link = soup.select('#articleWrap > div > div > img')
                try:
                    embed = discord.Embed(color=0xff4040, title="__" + links[int(str(message.content).split('동선 ')[1]) - 1].text.replace('◇ ', '') + "__")
                    linkr = str(link[int(str(message.content).split('동선 ')[1]) - 1]).split('src="')[1].split('">')[0]
                    embed.set_image(url='https:' + linkr)
                    await message.channel.send(embed=embed)
                except:
                    embed = discord.Embed(color=0xff4040, title="__확진자 동선 목록__")
                    #print(links[0].text)
                    rcount = 0
                    for i in links:
                        rcount = rcount + 1
                        if rcount == 1:
                            embed.add_field(name="목록번호 : " + str(rcount), value=i.text.replace('◇ ', ''), inline=True)
                        elif rcount % 3 == 0:
                            embed.add_field(name="목록번호 : " + str(rcount), value=i.text.replace('◇ ', ''), inline=False)
                        else:
                            embed.add_field(name="목록번호 : " + str(rcount), value=i.text.replace('◇ ', ''), inline=True)
                    embed.set_footer(text='!코로나 동선 목록번호 로 확인할수있습니다.')
                    sent = await message.channel.send(embed=embed)
                    await asyncio.sleep(30)
                    await sent.delete()
                    await message.delete()

            if str(message.content).split(' ')[1]==("지역"):
                rlist=["서울","부산","대구","인천","광주","대전","울산","경기","강원","충북","충남","전북","전남","경북","경남","제주","세종"]
                if str(message.content).split('지역 ')[1] in rlist:
                    request = requests.get('http://ncov.mohw.go.kr/bdBoardList_Real.do?brdId=1&brdGubun=13&ncvContSeq=&contSeq=&board_id=&gubun=')
                    crequest = requests.get('http://ncov.mohw.go.kr/bdBoardList_Real.do')
                    chtml = crequest.text
                    csoup = BeautifulSoup(chtml, 'html.parser')
                    request.encoding = 'utf-8'
                    html = request.text
                    soup = BeautifulSoup(html, 'html.parser')
                    links = html.split('<tbody>')[1].split('</tbody>')[0]
                    cdate = csoup.select('#content > div > div.bv_content > div > p')
                    cdater = str(cdate[1].text).split('보고(')[1].replace(')', '')
                    result=links.split('<th scope="row">'+str(message.content).split('지역 ')[1]+'</td>')[1].split('</tr>')[0]
                    embed = discord.Embed(color=0xff4040, title="__"+str(message.content).split('지역 ')[1] +" 코로나바이러스-19 현황 입니다.__")
                    embed.add_field(name="확진", value=result.split('<td headers="status_con s_type1" class="number">')[1].split('</td>')[0] +"명")
                    embed.add_field(name="완치", value=result.split('<td headers="status_con s_type2" class="number">')[1].split('</td>')[0] +"명")
                    embed.add_field(name="사망", value=result.split('<td headers="status_con s_type3" class="number">')[1].split('</td>')[0] +"명")
                    if str(message.content).split('지역 ')[1] == "경기":
                        embed.add_field(name="검사중", value=result.split('<td headers="status_con s_type4" class="number">')[1].split('</td>')[0] + "명")
                    else:
                        embed.add_field(name="검사중", value=result.split('<td headers="status_test s_type4" class="number">')[1].split('</td>')[0] +"명")
                    embed.add_field(name="결과음성", value=result.split('<td headers="status_test s_type5" class="number">')[1].split('</td>')[0] +"명")
                    embed.add_field(name="합계", value=result.split('<td class="number">')[1].split('</td>')[0] +"명")
                    embed.set_footer(text="질병관리본부 " + cdater)
                    await message.channel.send(embed=embed)


                else:
                    embed = discord.Embed(color=0xff4040, title="__다음 지역중 하나를 입력해주세요.__")
                    embed.add_field(name="지역목록 : ", value="서울, 부산, 대구, 인천, 광주, 대전\n울산, 경기, 강원, 충북, 전북, 전남\n경북, 경남, 제주, 세종", inline=True)
                    await message.channel.send(embed=embed)

        except IndexError:
            embed = discord.Embed(color=0xff4040, title="__국내 코로나바이러스-19 현황 입니다.__")
            cdate = soup.select('#content > div > div.bv_content > div > p')
            cdater = str(cdate[1].text).split('보고(')[1].replace(')','')
            embed.add_field(name="**현황**", value=str(links[0].text).replace('(','').replace(')',' :')+
                                             "\n"+str(links[1].text).replace('(','').replace(')',' :')+
                                             "\n"+str(links[2].text).replace('(','').replace(')',' :')+
                                             "\n"+str(links[3].text).replace('(','').replace(')',' :'))
            embed.set_footer(text="질병관리본부 "+cdater)
            await message.channel.send(embed=embed)

        file = openpyxl.load_workbook("statement.xlsx")
        sheet = file.active
        i = 1
        while True:
            if sheet["A" + str(i)].value == str(message.guild.id):
                sheet["B" + str(i)].value = sheet["B" + str(i)].value + 1
                file.save("statement.xlsx")
                break

            if sheet["A" + str(i)].value == None:
                sheet["A" + str(i)].value = str(message.guild.id)
                sheet["B" + str(i)].value = 1
                file.save("statement.xlsx")
                break
            i+=1
    
    #이 명령을 사용하기 위해선 nchannel.xlsx 파일이 필요합니다.
    if message.content.startswith("!공지"): #해당 채널에서 만 사용가능 
        if message.channel.id == 681524242076270670:
            file = openpyxl.load_workbook("nchannel.xlsx")
            sheet = file.active
            i=1
            while True:
                if sheet["A" + str(i)].value == None:
                    file.save("nchannel.xlsx")
                    break;
                if sheet["B" + str(i)].value == 1:
                    embed = discord.Embed(color=0xff4040, title="__실시간 공지__")
                    embed.add_field(name="코로나 알림 포럼", value=str(message.content).split('!공지 ')[1])
                    embed.set_footer(text="해당메세지는 코로나 알림 포럼 https://discord.gg/nBUGWPf에서 발송되었습니다.")
                    try:
                        await client.get_channel(int(sheet["A" + str(i)].value)).send(embed=embed)
                    except AttributeError:
                        pass;
                i+=1
    #이 아래부분은 네이버 데이터랩 api를 활용하여 제작하였습니다.
    if message.content.startswith("!이미지제작"):
        if message.channel.id == 848518003657408513: #해당 채널에서 만 사용가능 
            client_id = "네이버 api id"
            client_secret = "네이버 api secret"
            url = "https://openapi.naver.com/v1/datalab/search";
            now = time.gmtime(time.time())
            ndate=str(now.tm_year) + "-0" + str(now.tm_mon) + "-" + str(now.tm_mday-1)
            body = "{\"startDate\":\"2020-01-19\",\"endDate\":\""+ndate+"\",\"timeUnit\":\"date\",\"keywordGroups\":[{\"groupName\":\"코로나\",\"keywords\":[\"코로나\",\"우한폐렴\",\"개학연기\",\"대구\",\"사망\",\"완치\"]}]}";
            request = urllib.request.Request(url)
            request.add_header("X-Naver-Client-Id", client_id)
            request.add_header("X-Naver-Client-Secret", client_secret)
            request.add_header("Content-Type", "application/json")
            response = urllib.request.urlopen(request, data=body.encode("utf-8"))
            rescode = response.getcode()
            if (rescode == 200):
                response_body = response.read()
                result = json.loads(response_body)
                ratio = [each['ratio'] for each in result['results'][0]['data']]
                date = [each['period'] for each in result['results'][0]['data']]
                print(date + ratio)
                plt.figure(figsize=(18, 6))
                plt.plot(date, ratio)
                plt.grid()
                plt.xticks(rotation=90)
                plt.savefig('image.png')
                await message.channel.send(file=discord.File('image.png'))
                #print(response_body.decode('utf-8'))
            else:
                print("Error Code:" + rescode)
    
    #이미지 설정은 !코로나 타임라인 명령어 사진을 바꿉니다. 
    if message.content.startswith("!이미지설정"):
        if message.channel.id == 848518003657408513:
            f = open('tlurl.txt', 'w')
            f.write(str(message.content).split(' ')[1])
            f.close()




        #https://www.yna.co.kr/view/AKR20200221051400505?section=safe/news



client.run("ODQxNDc4MjkwMjExODY0NTg2.YJnVug.dJV161SXtXw6la-QeOwJZhtUN5g")
